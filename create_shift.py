import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_shift_management_table(filename="shift_management.xlsx"):
    # ワークブックとワークシートの作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シフト管理表"
    
    # グリッド線を表示する
    ws.views.sheetView.showGridLines = True
    
    # スタイル定義
    font_title = Font(name="Meiryo UI", size=16, bold=True, color="1B365D")
    font_header = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Meiryo UI", size=11)
    font_summary = Font(name="Meiryo UI", size=11, bold=True)
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_weekend_sat = PatternFill(start_color="E8F1F5", end_color="E8F1F5", fill_type="solid")
    fill_weekend_sun = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_summary = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    double_side = Side(border_style="double", color="000000")
    
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. タイトル
    ws["A1"] = "2026年9月 シフト管理表"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 40

    # データ構造の定義
    days = list(range(1, 31)) # 1日〜30日
    # 2026年9月1日は火曜日 (0:月, 1:火, 2:水, 3:木, 4:金, 5:土, 6:日)
    wdays = ["火", "水", "木", "金", "土", "日", "月"]
    staff_list = ["山田 太郎", "佐藤 花子", "鈴木 一郎", "高橋 次郎", "田中 節子"]

    # 2. ヘッダー作成 (行3: 日付, 行4: 曜日)
    ws.cell(row=3, column=1, value="氏名").font = font_header
    ws.cell(row=3, column=1).fill = fill_header
    ws.cell(row=3, column=1).alignment = align_center
    ws.cell(row=3, column=1).border = border_all
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    
    for idx, day in enumerate(days):
        col = idx + 2
        wday = wdays[idx % 7]
        
        # 日付セル
        cell_day = ws.cell(row=3, column=col, value=f"{day}日")
        cell_day.font = font_header
        cell_day.fill = fill_header
        cell_day.alignment = align_center
        cell_day.border = border_all
        
        # 曜日セル
        cell_wday = ws.cell(row=4, column=col, value=wday)
        cell_wday.font = font_header
        cell_wday.fill = fill_header
        cell_wday.alignment = align_center
        cell_wday.border = border_all
        
    # 集計ヘッダー
    summary_cols = ["出勤日数", "公休日数", "有給日数"]
    start_sum_col = len(days) + 2
    for idx, s_title in enumerate(summary_cols):
        col = start_sum_col + idx
        cell_sum = ws.cell(row=3, column=col, value=s_title)
        cell_sum.font = font_header
        cell_sum.fill = fill_header
        cell_sum.alignment = align_center
        cell_sum.border = border_all
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    # 3. サンプルデータ入力とスタイル適用
    shift_patterns = [
        ["早", "遅", "日", "公", "早", "遅", "公"] * 5,
        ["遅", "日", "公", "早", "遅", "公", "日"] * 5,
        ["日", "公", "早", "遅", "公", "日", "有"] * 5,
        ["公", "早", "遅", "公", "日", "早", "遅"] * 5,
        ["早", "公", "遅", "日", "早", "公", "日"] * 5,
    ]

    current_row = 5
    for i, staff in enumerate(staff_list):
        cell_staff = ws.cell(row=current_row, column=1, value=staff)
        cell_staff.font = font_body
        cell_staff.alignment = align_left
        cell_staff.border = border_all
        
        pattern = shift_patterns[i % len(shift_patterns)]
        
        for idx in range(len(days)):
            col = idx + 2
            wday = wdays[idx % 7]
            shift_val = pattern[idx]
            
            cell_shift = ws.cell(row=current_row, column=col, value=shift_val)
            cell_shift.font = font_body
            cell_shift.alignment = align_center
            cell_shift.border = border_all
            
            if wday == "土":
                cell_shift.fill = fill_weekend_sat
            elif wday == "日":
                cell_shift.fill = fill_weekend_sun
                
        col_letter_start = get_column_letter(2)
        col_letter_end = get_column_letter(len(days) + 1)
        
        f_work = f'=COUNTIF({col_letter_start}{current_row}:{col_letter_end}{current_row}, "早") + COUNTIF({col_letter_start}{current_row}:{col_letter_end}{current_row}, "日") + COUNTIF({col_letter_start}{current_row}:{col_letter_end}{current_row}, "遅")'
        f_off = f'=COUNTIF({col_letter_start}{current_row}:{col_letter_end}{current_row}, "公")'
        f_paid = f'=COUNTIF({col_letter_start}{current_row}:{col_letter_end}{current_row}, "有")'
        
        formulas = [f_work, f_off, f_paid]
        for idx, formula in enumerate(formulas):
            col = start_sum_col + idx
            cell_f = ws.cell(row=current_row, column=col, value=formula)
            cell_f.font = font_body
            cell_f.alignment = align_right
            cell_f.border = border_all
            
        ws.row_dimensions[current_row].height = 24
        current_row += 1

    # 4. 縦方向の集計行
    summary_rows = [("早番人数", "早"), ("日勤人数", "日"), ("遅番人数", "遅")]
    for s_idx, (label, symbol) in enumerate(summary_rows):
        cell_label = ws.cell(row=current_row, column=1, value=label)
        cell_label.font = font_summary
        cell_label.alignment = align_left
        cell_label.fill = fill_summary
        cell_label.border = border_all
        
        for idx in range(len(days)):
            col = idx + 2
            col_letter = get_column_letter(col)
            formula = f'=COUNTIF({col_letter}5:{col_letter}{current_row-1}, "{symbol}")'
            
            cell_count = ws.cell(row=current_row, column=col, value=formula)
            cell_count.font = font_body
            cell_count.alignment = align_right
            cell_count.fill = fill_summary
            cell_count.border = border_all
            
        for c in range(start_sum_col, start_sum_col + len(summary_cols)):
            ws.cell(row=current_row, column=c).border = border_all
            ws.cell(row=current_row, column=c).fill = fill_summary
            
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # 5. 列幅の自動調整
    ws.column_dimensions["A"].width = 16
    for idx in range(len(days)):
        col_letter = get_column_letter(idx + 2)
        ws.column_dimensions[col_letter].width = 6
    for idx in range(len(summary_cols)):
        col_letter = get_column_letter(start_sum_col + idx)
        ws.column_dimensions[col_letter].width = 12

    # 6. 保存
    wb.save(filename)
    print(f"'{filename}' が正常に作成されました。")

if __name__ == "__main__":
    create_shift_management_table()